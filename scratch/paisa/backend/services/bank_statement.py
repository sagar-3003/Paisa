import io
import json
import pdfplumber
import openpyxl
from services.llm import call_llm

def parse_bank_row(row):
    """Attempt heuristically to pull out date, desc, dr, cr, bal from row of text. Highly dependent on actual bank statements."""
    if len(row) < 4:
        return None
    
    # Generic simple extraction assuming 1=date, 2=desc, 3=type, 4/5=dr/cr
    try:
        return {
            "date": str(row[0] or ""),
            "description": str(row[1] or ""),
            "debit": float(row[2].replace(",", "")) if row[2] else 0.0,
            "credit": float(row[3].replace(",", "")) if len(row) > 3 and row[3] else 0.0,
        }
    except ValueError:
        return None

def parse_bank_statement_pdf(pdf_bytes: bytes) -> list[dict]:
    """Extract transactions from bank statement PDF."""
    transactions = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table[1:]: # Skip headers assuming 1st row
                        txn = parse_bank_row(row)
                        if txn and (txn['debit'] or txn['credit']):
                            transactions.append(txn)
    except Exception as e:
        print(f"Error extracting PDF tables: {e}")
    return transactions

def parse_bank_statement_excel(excel_bytes: bytes) -> list[dict]:
    transactions = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        headers = [str(c.value).lower() for c in ws[0]] if hasattr(ws, '__getitem__') else [str(c.value).lower() for c in ws[1]] # Support 0 or 1 index base
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            row_dict = dict(zip(headers, row))
            date_val = row_dict.get("date") or row_dict.get("txn date")
            desc_val = row_dict.get("description") or row_dict.get("narration") or row_dict.get("particulars")
            if not date_val or not desc_val:
                continue
                
            debit_val = str(row_dict.get("debit") or row_dict.get("withdrawal") or "0").replace(",", "")
            credit_val = str(row_dict.get("credit") or row_dict.get("deposit") or "0").replace(",", "")

            try:
                transactions.append({
                    "date": str(date_val),
                    "description": str(desc_val),
                    "debit": float(debit_val),
                    "credit": float(credit_val),
                    "balance": float(str(row_dict.get("balance") or "0").replace(",", "")),
                })
            except ValueError:
                continue
    except Exception as e:
        print(f"Error parsing Excel: {e}")
    return transactions

BANK_CATEGORIZE_PROMPT = """
Categorize each bank transaction for an Indian business.
Return ONLY a JSON array. Each item: {{"id": N, "category": "...", "ledger": "...", "tds_applicable": bool}}

Categories: salary, rent, vendor_payment, customer_receipt, gst_payment,
            tds_payment, bank_charges, loan_emi, utility, misc_expense, misc_income

Transactions:
{transactions}
"""

async def categorize_transactions(transactions: list[dict]) -> list[dict]:
    """Sends batches of transactions to LLM for auto-categorization."""
    if not transactions:
        return []
        
    formatted_txns = []
    for i, t in enumerate(transactions):
        formatted_txns.append(f"[{i}] Date: {t['date']}, Desc: {t['description']}, Dr: {t['debit']}, Cr: {t['credit']}")
        
    prompt = BANK_CATEGORIZE_PROMPT.format(transactions="\n".join(formatted_txns[:20])) # Limit 20 per batch
    
    response = await call_llm(prompt)
    try:
        if "```json" in response:
             response = response.split("```json")[1].split("```")[0].strip()
        elif "```" in response:
             response = response.split("```")[1].split("```")[0].strip()
        categories = json.loads(response)
        
        # Merge back
        for i, cat in enumerate(categories):
            if i < len(transactions):
                transactions[i].update(cat)
                
        return transactions
    except Exception as e:
        print(f"LLM Categorization failed: {e}")
        return transactions
